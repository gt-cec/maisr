from flask import Flask, request, jsonify, session
from flask import render_template

# optionally import SimConnect (for MSFS)
try:
    from SimConnect import *
    USING_SIMCONNECT = True
except ImportError:
    USING_SIMCONNECT = False

import subprocess
import multiprocessing
import os, psutil
import logging
import math, random, time
import socket, struct, pickle
import json
import numpy as np

# optionally import Python Excel
try:
    from openpyxl import Workbook, load_workbook
    USING_EXCEL = True
except ImportError:
    USING_EXCEL = False

app = Flask(__name__)

# Creating simconnection
try:
    sm = SimConnect()
    aq = AircraftRequests(sm, _time=10)
except:
    sm = None
    aq = None

ships_process = None  # the C++ exe for spawning the ships
# init center surveillance area at 33.0, -118.0
shift_north = 0.00
shift_west = 0.02
scale = 86.4* 1/10 # lat long scale of the ship group
dest_lat = None
dest_long = None
old_lat = 0
old_long = 0
# init center surveillance area at 33.0, -118.0

# aircraft position
ac1_init = True
ac1_lat = 33.0
ac1_long = -118.0
ac1_old_lat = 33.0
ac1_old_long = -118.0
ac1_alt = 2000
ac1_heading = 0.01
ac1_old_heading = 0.01
ac1_Xpx = 0.0
ac1_Ypx = 0.0
ac1_headingPx = 0.0

ac2_init = True
ac2_lat = 33.0
ac2_long = -118.0
ac2_old_lat = 33.0
ac2_old_long = -118.0
ac2_alt = 2000
ac2_heading = 0.01
ac2_old_heading = 0.01
ac2_Xpx = 0.0
ac2_Ypx = 0.0
ac2_headingPx = 0.0

aircraft2Waypoint = [0, 0]  # RL: the current waypoint of the second aircraft

# gameWidth is 100NM ~ 1.44 lat long degrees 
# 1.44 degree * 60 min/degree = 86.4 minutes 
ac_scale = 86.4 * 1/10 # scale down to 10Nm real world width
allMouseData = []
phaseData = []
clickData = []
alertData = []
random.seed(time.time())
rand = random.randint(1000, 9999)
app.id = None

study_scenario = 1
study_sequence = 1
study_config = "A0"
study_survey = "none"

target_rl_step = 0
current_rl_step = 0
current_rl_observation = {
    "current step": 0,
    "observation": {
        "agent1 pos": [0, 0],
        "agent2 pos": [0, 0],
        "threat pos": [],
        "threat class": []
    }
}
reset_env = False
env_is_reset = True

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/control")
def control():
    return render_template("controller.html")

def pxToLat(pxY):
    return 33 - ((pxY-0.5) * ac_scale) / 60.0 

def pxToLong(pxX):
    return -118 + ((pxX-0.5) * ac_scale) / 60.0 

def latToPx(latY):
    return 0.5 - (latY - 33) * 60.0 / ac_scale

def longToPx(longX):
    return (((longX + 118) * 60.0 )/ ac_scale) + 0.5

# RL Step(), gets or increments the current RL step for the client to know to step
@app.route("/rl-step", methods=["GET"])
def rl_step_get():
    return jsonify({"current step": current_rl_step, "target step": target_rl_step, "aircraft 2 dest x": aircraft2Waypoint[0], "aircraft 2 dest y": aircraft2Waypoint[1]})

@app.route("/rl-step", methods=["POST"])
def rl_step_post():
    global target_rl_step, aircraft2Waypoint
    if "source" not in request.json:
        return jsonify({"result": "failure, missing 'source'"})
    elif request.json["source"] == "rl":
        # take action at this step
        aircraft2Waypoint = request.json["action"]
        target_rl_step += 10
    return jsonify({"result": "success", "current step": current_rl_step, "target step": target_rl_step, "aircraft 2 dest x": aircraft2Waypoint[0], "aircraft 2 dest y": aircraft2Waypoint[1]})

# Gets or sets the current state observation
@app.route("/rl-observation", methods=["GET"])
def rl_step_observation_get():
    current_rl_observation["target step"] = target_rl_step
    return jsonify(current_rl_observation)  # dict of form {"current step": 123, "observation": {...}}

@app.route("/rl-observation", methods=["POST"])
def rl_step_observation_post():
    global current_rl_observation
    if "observation" not in request.json:
        raise ValueError("Missing 'observation' key-value pair in the rl-observation POST request!")
    if "current step" not in request.json:
        raise ValueError("Missing 'current step' key-value pair in the rl-observation POST request!")
    current_rl_observation["observation"] = request.json["observation"]  # dict of form {"current step": 123, "observation": {...}}
    current_rl_observation["current step"] = request.json["current step"]
    current_rl_observation["target step"] = target_rl_step
    current_rl_observation["reward"] = request.json["reward"]
    current_rl_observation["terminated"] = request.json["terminated"]
    current_rl_observation["truncated"] = False
    return jsonify({"result": "success"})

# RL Reset(), reset the game
@app.route("/rl-reset", methods=["GET"])
def rl_reset_get():
    print("GET RESET")
    return jsonify({"reset env": reset_env, "env is reset": env_is_reset})

@app.route("/rl-reset", methods=["POST"])
def rl_reset_post():
    print("POST RESET")
    global reset_env, env_is_reset, current_rl_step, target_rl_step
    if "command" not in request.json:
        return jsonify({"result": "failure, need 'command' parameter"})
    elif request.json["command"] == "init reset":
        reset_env = True
        env_is_reset = False
    elif request.json["command"] == "end reset":
        reset_env = False
        env_is_reset = True
        current_rl_step = 0
        target_rl_step = 0
    else:
        return jsonify({"result": "failure, missing command"})
    return jsonify({"result": "success"})

@app.route("/aircraft", methods=["GET"])
def update_aircraft():
    if not USING_SIMCONNECT:
        return jsonify(
            {
                "aircraft 1 x": random.random(), 
                "aircraft 1 y": random.random(), 
                "aircraft 1 heading": random.random() * 3.14,
                "aircraft 2 x": random.random(), 
                "aircraft 2 y": random.random(), 
                "aircraft 2 heading": random.random() * 3.14
            }
        )
    
    global ac1_init, ac1_lat, ac1_long, ac1_heading, ac1_old_lat, ac1_old_long, ac1_old_heading, sm, aq
    # initialization hover
    if (ac1_init):
        aq.set("PLANE_LATITUDE", ac1_lat)
        aq.set("PLANE_LONGITUDE", ac1_long)
        aq.set("PLANE_ALTITUDE", ac1_alt)
        aq.set("PLANE_HEADING_DEGREES_TRUE", ac1_heading)#data.get("heading"))
        aq.set("PLANE_PITCH_DEGREES", 0.0)
        aq.set("PLANE_BANK_DEGREES", 0.0)
        
    # remove temp stabilization when MATLAB implements
    # sometimes SimConnect breaks and throws an OS Error, so we are saving the current lat/long when it works (or sending the last one)
    try:
        ac1_lat = aq.get("PLANE_LATITUDE")
        ac1_long = aq.get("PLANE_LONGITUDE")
        ac1_heading = aq.get("PLANE_HEADING_DEGREES_TRUE")
    except:
        logging.info("SimConnect Error in /var")
        try:
            sm = SimConnect()
            aq = AircraftRequests(sm, _time=10)
        except:
            pass
        pass

    if (ac1_lat != None): ac1_old_lat = ac1_lat
    if (ac1_long != None): ac1_old_long = ac1_long
    if (ac1_heading != None): ac1_old_heading = ac1_heading

    ac1_Xpx = longToPx(ac1_old_long)
    ac1_Ypx = latToPx(ac1_old_lat)
    ac1_headingPx = float(ac1_old_heading)

    return_dict = {"x": ac1_Xpx,
                   "y": ac1_Ypx,
                   "heading": ac1_headingPx}

    return jsonify(return_dict) 


# spawns the ships given waypoints
@app.route("/ships", methods=['POST'])
def spawn_ships():
    # assumes the 'ships' param is structured as:
    #
    # [
    #   [  # ship
    #     [ID, targetClass, threatClass]
    #     [x, y, s],  # waypoint
    #     [x, y, s],
    #     [x, y, s],
    #     ...
    #   ],
    #   ...
    # ]

    if not USING_SIMCONNECT:
        return ""
    
    data = request.get_json()
    ships = data.get('ships')
    numShips = len(ships)
    shipstring = ""

    # create blank array with MATLABarraysize=10 ship elements
    matlabObsArrSize = 20
    shipObstacles = []
    shipID = [0.0 for _ in range(matlabObsArrSize)]
    shipTargetClass = [0.0 for _ in range(matlabObsArrSize)]
    shipThreatClass = [0.0 for _ in range(matlabObsArrSize)]
    shipLong = [0.0 for _ in range(matlabObsArrSize)]
    shipLat = [0.0 for _ in range(matlabObsArrSize)]

    for i in range(numShips):
        ship = ships[i]
        shipID[i % matlabObsArrSize] =  float(ship[0][0])
        shipTargetClass[i % matlabObsArrSize]  = float(ship[0][1])
        shipThreatClass[i % matlabObsArrSize]  = float(ship[0][2])
        shipLong[i % matlabObsArrSize]  = pxToLong(ship[1][0])
        shipLat[i % matlabObsArrSize]  = pxToLat(ship[1][1])
        for waypoint in ship[1:]:
            shipstring += str(waypoint[0]) + "-" + str(waypoint[1]) + "-" + str(waypoint[2]) + ","
        shipstring += ";"
    
    with open("./waypoints.txt", "w") as f:
        f.write(str(shift_north) + "\n")  # shift north (slight changes should be on the order of 0.01)
        f.write(str(shift_west) + "\n")  # shift west (slight changes should be on the order of 0.01)
        f.write(str(scale) + "\n")  # lat long scale (group size)
        f.write(shipstring)

    # send ship info and spawn point to matlab for CBF obstacle creation
    shipObstacles = [shipID, shipLat, shipLong, shipTargetClass, shipThreatClass]
    matlab_update_ship(shipObstacles)

    global ships_process
    # delete the ships if they are already active
    if ships_process is not None:
        # the generated process spawns the AIObjects.exe, which is a separate process, so we kill all children of the spawned process
        # something simple like process.terminate() does not kill AIObjects.exe
        parent = psutil.Process(ships_process.pid)
        for child in parent.children(recursive=True):  # or parent.children() for recursive=False
            child.kill()
        parent.kill()      

    # create the process
    ships_process = multiprocessing.Process(target=start_ships, args=(shipstring,))
    ships_process.start()
    return ""

def start_ships(shipstring=""):
    if USING_SIMCONNECT:
        with open("./waypoints.txt", "rb") as f:
            subprocess.run([".\AIObjects.exe"], stdin=f, text=True, shell=False)
        print("\n done!")
    return
    
def matlab_update_ship(shipObs):
    MATLAB_IP = '127.0.0.1'
    MATLAB_PORT_OBS_ID = 8082
    MATLAB_PORT_OBS_LAT = 8084
    MATLAB_PORT_OBS_LONG = 8086
    MATLAB_PORT_OBS_TARGET_CLASS = 7088
    MATLAB_PORT_OBS_THREAT_CLASS = 7080
    
    # format obstacle info
    IDs = shipObs[0]
    ship_lats = shipObs[1]
    ship_longs = shipObs[2]
    ship_target_class = shipObs[3]
    ship_threat_class = shipObs[4]

    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)

    for i in range(len(IDs)):
        # send the ship obstacle info
        s.sendto(struct.pack('>d', IDs[i]), (MATLAB_IP, MATLAB_PORT_OBS_ID ))
        s.sendto(struct.pack('>d', ship_lats[i]), (MATLAB_IP, MATLAB_PORT_OBS_LAT))
        s.sendto(struct.pack('>d', ship_longs[i]), (MATLAB_IP, MATLAB_PORT_OBS_LONG))
        s.sendto(struct.pack('>d', ship_target_class[i]), (MATLAB_IP, MATLAB_PORT_OBS_TARGET_CLASS ))
        s.sendto(struct.pack('>d', ship_threat_class[i]), (MATLAB_IP, MATLAB_PORT_OBS_THREAT_CLASS ))

    return "success"


@app.route('/run-executable', methods=['POST'])
def run_executable():
    dataArr = request.get_json()
    
    for data in dataArr:
        id = data.get('id')
        tClass = data.get('class')
        speed = data.get('speed')
        x = data.get('x')
        y = data.get('y')
        wpx = data.get('wpx')
        wpy = data.get('wpy')
        print(f"ID: {id}. Class: {tClass}. x,y: ({x},{y}). wpx,wpy: ({wpx},{wpy}). Speed: {speed}")
    return "success"

# clientside configuration route, client reads this to determine what configuration to show
@app.route("/config", methods=["GET"])
def get_config():
    return jsonify({"config": study_config, "scenario": study_scenario, "sequence": study_sequence, "survey": study_survey})

@app.route("/config", methods=["POST"])
def post_config():
    data = request.get_json()

    survey = data.get("survey")
    sequence = int(data.get("sequence"))
    scenario = int(data.get("scenario"))

    square_order = ["A0", "B0", "B3", "A1", "A3", "B1", "B2", "A2",
                    "B0", "A1", "A0", "B1", "B3", "A2", "A3", "B2",
                    "A1", "B1", "B0", "A2", "A0", "B2", "B3", "A3",
                    "B1", "A2", "A1", "B2", "B0", "A3", "A0", "B3",
                    "A2", "B2", "B1", "A3", "A1", "B3", "B0", "A0",
                    "B2", "A3", "A2", "B3", "B1", "A0", "A1", "B0",
                    "A3", "B3", "B2", "A0", "A2", "B0", "B1", "A1",
                    "B3", "A0", "A3", "B0", "B2", "A1", "A2", "B1"]
    if sequence > 8:
        return "sequence too high, capped at 8"
    if scenario > 8:
        return "scenario too high, capped at 8"
    if sequence < 1:
        return "sequence too low, min is 1"
    if scenario < 1:
        return "scenario too low, min is 1"
    if sequence * scenario > len(square_order):
        return "sequence x scenario is greater than max: " + str(len(square_order))
    global study_scenario, study_sequence, study_config, study_survey

    study_config = square_order[((sequence-1) * 8) + scenario - 1]
    study_scenario = scenario
    study_sequence = sequence
    study_survey = survey
    return "Server: updated scenario (" + str(scenario) + ") and sequence (" + str(sequence) + "), new config: " + str(study_config) + ", last survey: " + str(survey)

# logging route
@app.route("/log", methods=["POST"])
def log():
    #print("logging!", request.get_json())
    return ""

log = logging.getLogger('werkzeug')
log.setLevel(logging.ERROR)

@app.route("/saveId", methods=["POST"])
def saveId():
    data = request.get_json()
    app.id = data.get('userId')
    print(app.id)
    return ""

@app.route("/saveData", methods=["POST"])
def saveData():
    if not USING_EXCEL:
        return ""
    
    data = request.get_json()
    newData = {
        "phase": data.get('phase'),
        "score": data.get('score'),
        "numClicks": data.get('numClicks'),
        "time": data.get('time'),
        "unixTime": data.get('unixTime'),
        "denied": data.get('denied'),
        "numOptimized": data.get('numOptimized'),
        "wez_alerts": data.get('wez_alerts'),
        "tmv_alerts": data.get('tmv_alerts'),
        "osi_alerts": data.get('osi_alerts'),
        "utc_alerts": data.get('utc_alerts')
    }
    phaseData.append(newData)
    filename = f"{app.id}_{study_config}.pkl"
    txtFilename = f"{app.id}_{study_config}.txt"

    with open(filename, 'wb') as file:
        pickle.dump(allMouseData, file)
    with open(filename, 'ab') as file:
        pickle.dump(phaseData, file)
    with open(filename, 'ab') as file:
        pickle.dump(clickData, file)
    with open(filename, 'ab') as file:
        pickle.dump(alertData, file)
    text_content = "\n".join(json.dumps(element) if isinstance(element, dict) else element for element in alertData)
    with open(txtFilename,"w") as file:
        file.write(text_content)

    #reset data structures at current mission end
    if (data.get('phase') == 2):
        exFilename = f"{app.id}_{study_config}_alerts.xlsx"
        if not os.path.isfile(exFilename):
            wb = Workbook()
            ws = wb.active
            ws.append(['User ID'] + ['WEZ'] + ['Optimized Search'] + ['Unable to Comply'] + ['Too Many Vectors'] 
                      + ['WEZ'] + ['Optimized Search'] + ['Unable to Comply'] + ['Too Many Vectors'])
            wb.save(exFilename)
        wb = load_workbook(exFilename)
        ws = wb.active
        ws.append([app.id] + [phaseData[0]['wez_alerts']] + [phaseData[0]['osi_alerts']] + [phaseData[0]['utc_alerts']] + [phaseData[0]['tmv_alerts']]
                  + [phaseData[1]['wez_alerts']] + [phaseData[1]['osi_alerts']] + [phaseData[1]['utc_alerts']] + [phaseData[1]['tmv_alerts']])
        wb.save(exFilename)
        phaseData.clear()
        clickData.clear()
        allMouseData.clear()
        alertData.clear()
    return ""

@app.route('/receive-mouse-coordinates', methods=["POST"])
def receive_mouse_coordinates():
    data = request.json
    allMouseData.extend(data)
    return jsonify({})

@app.route('/receive-mouse-click', methods=["POST"])
def receive_mouse_click():
    data = request.json
    clickData.extend(data)
    return jsonify({})

@app.route('/receive-alert', methods=["POST"])
def receive_alert():
    data = request.json
    alertData.extend(data)
    return jsonify({})

# data for MATLAB route
@app.route("/current-destination", methods=["POST"])
# UDP update for MATLAB
def matlab_destination_update():
    MATLAB_IP = '127.0.0.1'
    MATLAB_PORT_LAT_MIN = 8080
    MATLAB_PORT_LONG_MIN = 8088

    global dest_lat, dest_long, ac1_init
    # on first destination, release aircraft from init hover
    ac1_init = False
    data = request.get_json()
    dest_x = float(data.get('aircraft 1 dest x'))
    dest_y = float(data.get('aircraft 1 dest y'))
    dest_long = pxToLong(dest_x )  # convert to long
    dest_lat = pxToLat(dest_y)  # convert to lat
 
    if (dest_lat is None or dest_long is None):
        return "failure"

    # parse the latitude and longitude
    latitude = float(dest_lat)
    longitude = float(dest_long)

    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)

    # send the location
    s.sendto(struct.pack('>f', latitude), (MATLAB_IP, MATLAB_PORT_LAT_MIN))
    s.sendto(struct.pack('>f', longitude), (MATLAB_IP, MATLAB_PORT_LONG_MIN))

    return "success"

@app.route("/cognitivestate", methods=["POST"])
def matlab_cognitivestate():
    MATLAB_IP = '127.0.0.1'
    MATLAB_PORT_COG_STATE = 7082

    data = request.get_json()

    data_cogstate = data.get("cogstate")
    cogstate = float(data_cogstate)
    print(cogstate)

    s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
    s.sendto(struct.pack('>d', cogstate), (MATLAB_IP, MATLAB_PORT_COG_STATE))

    return "success"


if __name__ == '__main__':
    app.run("0.0.0.0", port=100, debug=False)
